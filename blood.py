from tkinter import *
import tkinter.font as tkfont
import tkinter.messagebox
from openpyxl import Workbook
import os
from openpyxl import load_workbook
import smtplib
from tkcalendar import DateEntry

#------------mail sending part works only for gmail users----------------

path = r"logs.xlsx"
myaddress = "enter your email id"
password = "enter your password"


if os.path.isfile(path) == True:
        wb = load_workbook("logs.xlsx")
        ws = wb.worksheets[0]
        print("workbook exists")
else:
    wb = Workbook()
    ws = wb.active
    print('false')
    ws.append(["blood present"])
    ws.append (["NAME", "PHONE NO", "EMAIL ID", "BLOOD TYPE", "GENDER", "DATE OF BIRTH", "UID"]),
    ws.cell(row=1, column=11, value="blood given")
    #for index, title in enumerate(("unique id ", "receiver's name"), start = 11):
        #ws.cell(row=2, column=index, value=title)




def submit():
    ws.column_dimensions['B'].width = 20
    fname = name.get()
    fbt = blood_type.get()
    fphone = str(phone.get())
    fuid = str(uid.get())
    feid = str(eid.get())
    fdob = str(dob.get())
    fsex = str(r.get())
    if (fname==''or fphone=='' or fbt=='' or fuid=='' or feid=='' or fdob=='' or fsex==''):
        tkinter.messagebox.showerror("Error", "data missing")
    elif (len(fphone) != 13):
        tkinter.messagebox.showerror("Error", "Invalid phone number")
    else:
        result = tkinter.messagebox.askquestion("submit", "The data is \n" + fname + "\n" + fphone + "\n" + feid + "\n"+ fbt + "\n" +fsex+ "\n" +fdob+ "\n" + fuid)
        if result == 'yes':
            print('done')
            d = [fname, fphone, feid , fbt, fsex, fdob, fuid]
            ws.append(d)
            wb.save("logs.xlsx")
            clear()

            submit_label = Label(tk, text="DEAILS RECORDED" , font =('Times New Roman',  '15'), bg='#f5bdb3')
            submit_label.place(x=800, y=90)




        else:
            clear()



def clear():
    name.set('')
    blood_type.set('')
    phone.set('+91')
    uid.set('')
    eid.set('')
    dob.set('')



def mail():
    myadd = myaddress
    mypass = password
    fnuid = nuid.get()
    for cell in ws['G']:
        if cell.value == fnuid:
            r = cell.row
            ans = ws.cell(row=r, column=3)
            ans1 = ws.cell(row=r, column=4)
            new = str(ans.value)
            new1 = str(ans1.value)
            try:
                with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
                    smtp.ehlo()
                    smtp.starttls()
                    smtp.ehlo()
                    smtp.login(myadd, mypass)
                    subject = "BLOOD IN USE"
                    body = "Thank you for donating your blood of group "+new1+ " \nThis mail is to inform you that your blood is currently in use "
                    msg = f'subject: {subject}\n\n{body}'
                    smtp.sendmail(myadd, new, msg)
                    break
            except smtplib.SMTPAuthenticationError:
                tkinter.messagebox.showerror("ERROR" , "MAIL NOT DELIVERED!!!\nCheck your Login Credentials Properly.\n"
                      "If still the error persists follow the following steps given below:\n"
                      "login to gmail acc<<manage account<< security<< Less secure app access<< turn it on.")
                break
            except Exception:
                tkinter.messagebox.showerror("Error", "An unnatural exception occurred try checking your internet connection..restart the program")
                break
    else:
        tkinter.messagebox.showerror("Error", "INCORRECT UID")



tk = Tk()
tk.iconbitmap(r'favicon.ico')
tk.configure(bg='#f5bdb3')
tk.geometry("1000x600")
tk.title('Blood Bank')
data_font = tkfont.Font(size=15, weight='bold')
f = LabelFrame(tk, text='hey there')
f.pack(padx=5,pady=5)

#******var declaration**********
name =StringVar()
blood_type = StringVar()
phone = StringVar()
uid = StringVar()
nuid = StringVar()
rname = StringVar
eid  = StringVar()
dob = StringVar()
r = StringVar()


#**********heading**************
head_font = tkfont.Font(size=25, weight='bold')
head = Label(tk, text='BLOOD BANK ENTRY 🙂', bd=10, font = head_font  ).place(x=310, y=20)

#**********NAME**************
donor_name = Label(tk, text="DONOR'S NAME", font=data_font)
donor_name.place(x=70, y=150)
donor_entry_name = Entry(tk,textvariable=name, bd ='4', selectborderwidth='4',font=data_font)
donor_entry_name.place(x=250, y=150)

#**************PHONE NO*******************
donor_phone = Label(tk, text="PHONE NO",font=data_font)
donor_phone.place(x=80,y=200)
donor_phone_Entry = Entry(tk,textvariable=phone, bd ='4', selectborderwidth='4',font=data_font)
phone.set('+91')
donor_phone_Entry.place(x=250,y=200)

#*************blood type****************
donor_blood_type = Label(tk, text="BLOOD GROUP", font=data_font)
donor_blood_type.place(x=550, y=150)
l1 = [ 'A+','A-','B+','B-','AB+', 'AB-','O+', 'O-']
donor_entry_blood_type = OptionMenu(tk,blood_type,*l1)
donor_entry_blood_type.config(font=('Helvetica', 12),width=21)
blood_type.set('A+')
donor_entry_blood_type.place(x=720, y=150)

#****************unique id***************
donor_uid = Label(tk, text="UNIQUE ID", font=data_font)
donor_uid.place(x=599, y=200)
donor_entry_uid = Entry(tk,textvariable=uid, bd ='4', selectborderwidth='4',font=data_font)
donor_entry_uid.place(x=720, y=200)

#***************mail id****************
donor_uid = Label(tk, text="EMAIL ID", font=data_font)
donor_uid.place(x=80, y=250)
donor_entry_uid = Entry(tk,textvariable=eid, bd ='4', selectborderwidth='4',font=data_font)
donor_entry_uid.place(x=250, y=250)



#*************NEW unique id*****
donor_uid_new = Label(tk, text="UNIQUE ID", font=data_font)
donor_uid_new.place(x=80, y=450)
donor_entry_uid_new = Entry(tk,textvariable=nuid, bd ='4', selectborderwidth='4',font=data_font)
donor_entry_uid_new.place(x=200, y=450)

#************receiver name**********
receiver_name = Label(tk, text="RECEIVERS NAME",font=data_font)
receiver_name.place(x=530,y=450)
receiver_name_Entry = Entry(tk,textvariable=rname, bd ='4', selectborderwidth='4',font=data_font)
receiver_name_Entry.place(x=720,y=450)

#************age******************
age = Label(tk, text="AGE", font=data_font)
age.place(x=599, y=250)
age_entry = DateEntry(tk,  bd ='4',textvariable = dob , background= 'darkblue', foreground= 'white',year = 2000, font=data_font)
age_entry.place(x=720, y=250)

#*********gender**********
sex = Label(tk, text="GENDER" ,font=data_font)
sex.place(x=90, y=300)
male = Radiobutton(tk, text='MALE',value='male',variable = r, font=data_font)
male.place(x=250, y=300)
female = Radiobutton(tk, text='FEMALE',value='female',variable = r, font=data_font)
female.place(x=350, y=300)
r.set('male')


#***********all buttons**************
submit_button = Button(tk, text='SUBMIT',command=submit, padx = 50, font = data_font)
submit_button.place(x=100,y=350)
clear_button = Button(tk, text='CLEAR',command = clear,padx = 50, font = data_font)
clear_button.place(x=400,y=350)
exit_button = Button(tk, text=' EXIT',command = quit ,padx = 50, font = data_font)
exit_button.place(x=700,y=350)
mail_button = Button(tk, text='SEND MAIL',command= mail, padx = 50, font = data_font)
mail_button.place(x=400,y=500)






tk.mainloop()